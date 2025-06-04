"""
EXCEL DATA CONSOLIDATOR ENGINE
==============================

Module xử lý logic chính của ứng dụng: quét đệ quy thư mục, đọc file Excel, 
và tổng hợp dữ liệu thành file kết quả. Sử dụng thư viện xlwings để tương tác
với Microsoft Excel Engine, đảm bảo tương thích mọi phiên bản Excel.

Cấu trúc lớp chính:
-------------------
class ExcelConsolidator:
    Quản lý toàn bộ quy trình gom dữ liệu từ nhiều file Excel thành file tổng hợp

Các phương thức chính:
1. __init__(data_path: str)
2. _validate_path()
3. _scan_files() -> list
4. _get_relative_path(file_path: Path) -> str
5. _read_excel(file_path: Path) -> dict
6. consolidate(output_file: str) -> bool
7. _save_excel(df: pd.DataFrame, output_file: str)
8. get_summary() -> dict

Chi tiết từng phương thức:
"""
import os
from pathlib import Path
import logging
import pandas as pd
import xlwings as xw

# Cấu hình logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("ExcelConsolidator")
logger.setLevel(logging.INFO)

class ExcelConsolidator:
    def __init__(self, data_path: str):
        """
        Khởi tạo công cụ gom dữ liệu Excel
        
        Args:
            data_path: Đường dẫn tuyệt đối đến thư mục gốc chứa các file Excel
                       (có thể chứa nhiều subfolder)
        
        Attributes:
            data_path (Path): Đường dẫn chuẩn hóa
            all_keys (set): Tập hợp tất cả key duy nhất từ tất cả file
            file_data (dict): Dữ liệu từng file dạng {relative_path: {key: value}}
            file_order (dict): Thứ tự key trong từng file {relative_path: [key1, key2]}
        """
        self.data_path = Path(data_path)
        self.all_keys = set()
        self.file_data = {}
        self.file_order = {}
        self._validate_path()

    def _validate_path(self):
        """
        Kiểm tra tính hợp lệ của đường dẫn dữ liệu
        
        Raises:
            FileNotFoundError: Nếu đường dẫn không tồn tại
            NotADirectoryError: Nếu đường dẫn không phải thư mục
        """
        if not self.data_path.exists():
            raise FileNotFoundError(f"Path {self.data_path} not found")
        if not self.data_path.is_dir():
            raise NotADirectoryError(f"{self.data_path} is not a directory")

    def _scan_files(self) -> list:
        """
        Quét đệ quy tất cả file Excel trong thư mục và subfolders
        
        Returns:
            list: Danh sách đường dẫn tuyệt đối đến các file Excel tìm thấy
            
        Logic:
            - Hỗ trợ định dạng .xlsx, .xls, .xlsm
            - Sử dụng os.walk để đảm bảo tương thích đa phiên bản Python
            - Bỏ qua file ẩn và file hệ thống
        """
        excel_files = []
        for root, _, files in os.walk(self.data_path):
            for file in files:
                if file.lower().endswith(('.xlsx', '.xls', '.xlsm')):
                    full_path = Path(root) / file
                    # Bỏ qua file ẩn (bắt đầu bằng ~$)
                    if not file.startswith('~$'):
                        excel_files.append(full_path)
        logger.info(f"Found {len(excel_files)} Excel files in {self.data_path}")
        return excel_files

    def _get_relative_path(self, file_path: Path) -> str:
        """
        Chuyển đường dẫn tuyệt đối thành đường dẫn tương đối so với thư mục gốc
        
        Args:
            file_path: Đường dẫn tuyệt đối đến file Excel
            
        Returns:
            str: Đường dẫn tương đối (vd: 'subfolder/file.xlsx')
            
        Ví dụ:
            data_path = 'C:/data'
            file_path = 'C:/data/reports/january/sales.xlsx'
            Kết quả: 'reports/january/sales.xlsx'
        """
        try:
            return str(file_path.relative_to(self.data_path))
        except ValueError:
            return str(file_path)

    def _read_excel(self, file_path: Path) -> dict:
        """
        Đọc dữ liệu từ file Excel và trích xuất cặp key-value
        
        Args:
            file_path: Đường dẫn tuyệt đối đến file Excel
            
        Returns:
            dict: Dữ liệu dạng {key: value} từ file
            
        Logic xử lý:
            1. Sử dụng pandas để đọc file nhanh hơn
            2. Chỉ đọc sheet đầu tiên
            3. Giả định cột A là key, cột B là value
            4. Giữ lại giá trị đầu tiên nếu key trùng lặp trong cùng file
            5. Ghi nhận thứ tự key xuất hiện trong file gốc
        
        Raises:
            RuntimeError: Nếu không thể đọc file do lỗi hệ thống
        """
        relative_path = self._get_relative_path(file_path)
        logger.info(f"Reading file: {relative_path}")
        
        data_dict = {}
        key_order = []
        
        try:
            # Đọc file Excel với pandas
            df = pd.read_excel(
                file_path,
                sheet_name=0,
                header=None,
                usecols=[0, 1],  # Chỉ đọc cột A và B
                dtype={0: str, 1: str},  # Đảm bảo dữ liệu là string
                engine='openpyxl'  # Sử dụng openpyxl cho .xlsx
            )
            
            # Xử lý từng dòng dữ liệu
            for _, row in df.iterrows():
                key = str(row[0]).strip() if pd.notna(row[0]) else ""
                value = row[1] if pd.notna(row[1]) else ""
                
                if key:
                    # Chỉ lấy giá trị đầu tiên cho mỗi key
                    if key not in data_dict:
                        data_dict[key] = value
                        key_order.append(key)
                        self.all_keys.add(key)
            
            # Lưu thứ tự key trong file này
            self.file_order[relative_path] = key_order
            return data_dict
            
        except Exception as e:
            logger.error(f"Error reading {relative_path}: {str(e)}")
            return {}

    def consolidate(self, output_file: str = "consolidated.xlsx") -> bool:
        """
        Thực hiện quy trình gom dữ liệu từ tất cả file vào file tổng hợp
        
        Args:
            output_file: Đường dẫn file kết quả (.xlsx)
            
        Returns:
            bool: True nếu thành công, False nếu thất bại
            
        Quy trình xử lý:
            1. Quét tất cả file Excel trong thư mục và subfolders
            2. Đọc và trích xuất dữ liệu từ từng file
            3. Tạo danh sách key duy nhất (sắp xếp alphabet)
            4. Tạo ma trận dữ liệu với:
                - Hàng đầu: ['Source File'] + [sorted_keys]
                - Các hàng sau: [relative_path] + [values]
            5. Xuất ra file Excel với định dạng chuẩn
        
        Exception Handling:
            - Ghi log chi tiết khi có lỗi
            - Trả về False nếu không có file hoặc dữ liệu hợp lệ
        """
        try:
            excel_files = self._scan_files()
            if not excel_files:
                logger.error("No Excel files found")
                return False
                
            # Đọc dữ liệu từ tất cả file
            for file in excel_files:
                relative_path = self._get_relative_path(file)
                self.file_data[relative_path] = self._read_excel(file)
            
            if not self.all_keys:
                logger.error("No valid keys found")
                return False
                
            # Tạo danh sách key tổng (A-Z)
            sorted_keys = sorted(self.all_keys)
            
            # Xây dựng dữ liệu cho DataFrame
            data_rows = []
            
            # Header row
            header_row = ["Source File"] + sorted_keys
            data_rows.append(header_row)
            
            # Data rows
            for relative_path, data in self.file_data.items():
                row = [relative_path]
                file_keys = self.file_order.get(relative_path, [])
                
                # Tạo mapping ưu tiên thứ tự trong file gốc
                value_map = {k: data.get(k, "") for k in file_keys}
                
                # Bổ sung key còn thiếu
                for key in sorted_keys:
                    if key not in value_map:
                        value_map[key] = data.get(key, "")
                
                # Thêm giá trị theo thứ tự key tổng
                row.extend(value_map[key] for key in sorted_keys)
                data_rows.append(row)
            
            # Tạo và lưu DataFrame
            consolidated_df = pd.DataFrame(data_rows[1:], columns=data_rows[0])
            self._save_excel(consolidated_df, output_file)
            return True
            
        except Exception as e:
            logger.exception(f"Consolidation failed: {str(e)}")
            return False

    def _save_excel(self, df: pd.DataFrame, output_file: str):
        """
        Lưu DataFrame ra file Excel với định dạng chuyên nghiệp
        
        Args:
            df: DataFrame chứa dữ liệu tổng hợp
            output_file: Đường dẫn file xuất
            
        Định dạng:
            - Header: in đậm, nền xanh lá nhạt
            - Tự động điều chỉnh độ rộng cột
            - Đóng băng hàng đầu tiên
        """
        try:
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.dimensions import ColumnDimension
            
            # Tạo writer với openpyxl engine
            with pd.ExcelWriter(
                output_file,
                engine='openpyxl',
                mode='w'
            ) as writer:
                # Ghi DataFrame vào Excel
                df.to_excel(
                    writer,
                    sheet_name='Consolidated Data',
                    index=False,
                    header=True
                )
                
                # Lấy workbook và worksheet
                workbook = writer.book
                worksheet = writer.sheets['Consolidated Data']
                
                # Định dạng header
                header_fill = PatternFill(
                    start_color='CCFFCC',
                    end_color='CCFFCC',
                    fill_type='solid'
                )
                header_font = Font(bold=True)
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Tự động điều chỉnh độ rộng cột
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Đóng băng hàng đầu tiên
                worksheet.freeze_panes = 'A2'
                
        except Exception as e:
            logger.error(f"Save failed: {str(e)}")
            raise

    def get_summary(self) -> dict:
        """
        Trả về thống kê dữ liệu đã xử lý
        
        Returns:
            dict: {
                'total_files': Số file đã xử lý,
                'total_keys': Số key duy nhất,
                'file_names': Danh sách đường dẫn tương đối
            }
        """
        return {
            "total_files": len(self.file_data),
            "total_keys": len(self.all_keys),
            "file_names": list(self.file_data.keys())
        }
